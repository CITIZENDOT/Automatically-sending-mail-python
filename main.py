#Modules for excel file
import openpyxl
import datetime
#Modules for sending email
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders

wb = openpyxl.load_workbook("Appa.xlsx")
ws = wb["Sheet1"]

date_cell = ws.cell(5, 2)
today_date = datetime.date.today()
date_string = today_date.strftime("%d-%m-%Y")
date_cell.value = date_string
print("Tell me  what subject and topic covered from 10 am to 11 am: ")
subject_10am = input()
print("Tell me  what subject and topic covered from 4 pm to 5 pm: ")
subject_4pm = input()
subject_10_cell = ws.cell(5, 8)
subject_10_cell.value = subject_10am
subject_4_cell = ws.cell(5, 10)
subject_4_cell.value = subject_4pm

wb.save("Appa.xlsx")
print("File changes: Done")

fromaddr = input("Enter your email address: ")
password = input("Enter your password: ")
toaddr = input("Enter the 'to address' of mail and press Enter: ")
print("Preparing Mail attachments...")
msg = MIMEMultipart()
msg["From"] = fromaddr
msg["To"] = toaddr
msg["Subject"] = "Z. P. H. S, Arlabanda"
body = ""
msg.attach(MIMEText(body, "plain"))

filename = "Appa.xlsx"
attachment = open("Appa.xlsx", "rb")
p = MIMEBase("application", "octet-stream")

p.set_payload((attachment).read())

encoders.encode_base64(p)

p.add_header("Content-Disposition", "attachment; filename = %s"%filename)
msg.attach(p)

s = smtplib.SMTP("smtp.gmail.com", 587)

s.starttls()
print("Done")
print("Logging in...")
s.login(fromaddr, password)
print("Done")
text = msg.as_string()
print("Sending Mail...")
s.sendmail(fromaddr, toaddr, text)

s.quit()

print("Mail Sent")

